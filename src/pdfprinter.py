import os, sys
import win32print
from xlrd import open_workbook
from openpyxl import load_workbook
import subprocess
from tkinter import filedialog, Tk, Label, Button, LEFT, RIGHT, W, Message, StringVar, Toplevel, Listbox, ttk, Text, Checkbutton
from tkinter import *
import tkinter
from tkinter import ttk
import threading
from operator import itemgetter
import io
from time import asctime


def checkDependencies():
	pass
"""
def checkGhostScriptPath():
	try:
		pathKey = "HKEY_LOCAL_MACHINE\System\CurrentControlSet\Control\Session Manager\Environment"
		key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,"System\CurrentControlSet\Control\Session Manager\Environment",0,winreg.KEY_ALL_ACCESS)
		oldPath = winreg.QueryValueEx(key, "Path")
		dirsList = oldPath[0].rsplit(';')
		ghostscriptPath = r"C:\Program Files\gs\gs9.19"
		pathIsPresent = False
		for x in dirsList:
			if ghostscriptPath == x:
				pathIsPresent = True

		if not pathIsPresent:
			newPath =oldPath[0]+";"+ghostscriptPath
			winreg.SetValueEx(key,"Path",0,2,newPath)

		winreg.CloseKey(key)
	except:
		pass
"""

def determineColumnNumbers(worksheet, excelFormat):


	drawingColumn = 0

	if excelFormat == "xls":

		counter = 0
		for column in range(worksheet.ncols):
			value = str(worksheet.cell(0,column).value).rstrip()
			#print (value)
			if value == "partno":
				partColumn = column
			elif value == "drawingno":
				drawingColumn = column
			elif value == "revlevel":
				revColumn = column
			elif value == "parttype":
				gageColumn = column

	elif excelFormat == "xlsx":
		counter = 0
		for column in worksheet.columns:
			value = str(column[0].value).rstrip()
			#print (value)
			if value == "partno":
				partColumn = counter
			elif value == "drawingno":
				drawingColumn = counter
			elif value == "revlevel":
				revColumn = counter
			elif value == "parttype":
				gageColumn = counter
			counter += 1

	else:
		return None


	if drawingColumn:
		printOrderByFileFormat = "BURN"
	else:
		drawingColumn = 0
		gageColumn = 0
		printOrderByFileFormat = "WO"


	return (partColumn, drawingColumn, revColumn, gageColumn, printOrderByFileFormat)




def readExcel(fileToRead):
	partNumberList, drawingNOList, revLevelList, gageList = [],[],[],[]
	excelTempDict = {'PARTNO':'','DRAWING':'','REV':'','GAGE':''}
	dictList = []


	if fileToRead:
		if fileToRead.split(".")[1] == "xls" or fileToRead.split(".")[1] == "XLS":
			wb = open_workbook(str(fileToRead))
			ws = wb.sheets()[0]
			excelFormat = "xls"
			partColumn, drawingColumn, revColumn, gageColumn, printOrderByFileFormat = determineColumnNumbers(ws, excelFormat)
			for row in range(ws.nrows):
				excelTempDict = {}
				partnumber = str(ws.cell(row,partColumn).value).rstrip()
				if partnumber == 'None':
					break
				if not partnumber == 'partno':
					if (partnumber[-1] == 'F'):
						excelTempDict['PARTNO']=partnumber[:-1]
					elif(partnumber[-1] == 'S'):
						excelTempDict['PARTNO']=partnumber
					else:
						excelTempDict['PARTNO']=partnumber

					excelTempDict['DRAWING']=str(ws.cell(row,drawingColumn).value).rstrip()
					excelTempDict['REV']=str(ws.cell(row,revColumn).value).rstrip()
					excelTempDict['GAGE']=str(ws.cell(row,gageColumn).value).rstrip()
					dictList.append(excelTempDict)
					#print (excelTempDict)
					#print ()

		elif fileToRead.split(".")[1] == "xlsx" or fileToRead.split(".")[1] == "XLSX":

			ws = load_workbook(filename=str(fileToRead))['Sheet1']
			excelFormat = "xlsx"
			partColumn, drawingColumn, revColumn, gageColumn, printOrderByFileFormat = determineColumnNumbers(ws, excelFormat)
			for row in ws.rows:
				excelTempDict = {}
				partnumber = str(row[partColumn].value).rstrip()
				if partnumber == 'None':
					break

				if not partnumber == 'partno':
					if (partnumber[-1] == 'F'):
						excelTempDict['PARTNO']=partnumber[:-1]
					elif(partnumber[-1] == 'S'):
						excelTempDict['PARTNO']=partnumber
					else:
						excelTempDict['PARTNO']=partnumber

					excelTempDict['DRAWING']=str(row[drawingColumn].value).rstrip()
					excelTempDict['REV']=str(row[revColumn].value).rstrip()
					excelTempDict['GAGE']=str(row[gageColumn].value).rstrip()
					dictList.append(excelTempDict)
					#print (excelTempDict)
					#print ()

	return (dictList, printOrderByFileFormat)

def findPDFs(excelDictList):

	RELEASED_PDFS_LS = os.listdir("T:\RELEASED_FILES\CURRENT_PDF")
	#RELEASED_PDFS_LS = os.listdir("T:\ENGINEERING\FILE_SANDBOX\CADKEY TO PDF")

	foundParts, unfoundParts, wrongRevsion = [], [], []
	for dictx in excelDictList:
		for x in range(len(RELEASED_PDFS_LS)):

			if dictx['PARTNO'] in RELEASED_PDFS_LS[x]:
				if ('R'+dictx['REV'] in RELEASED_PDFS_LS[x]):
					foundParts.append("\""+RELEASED_PDFS_LS[x]+"\"")
					break
				else:
					wrongRevsion.append("\""+RELEASED_PDFS_LS[x]+"\"")
					break
			else:
				if(len(RELEASED_PDFS_LS)-1 == x):
					unfoundParts.append(dictx['PARTNO'])

	return [foundParts,unfoundParts, wrongRevsion]

def sortPartNumberList(excelInfo, printOrderByFileFormat):

	floatTempList, otherTempList = [], []
	dictsSortedByGage = sorted(excelInfo, key=itemgetter('GAGE'))
	previousGage = None
	gageSeperatedList = []
	allList = []


	if printOrderByFileFormat == "BURN":
		for dictx in dictsSortedByGage:
			if previousGage == None:
				previousGage = dictx['GAGE']

			if previousGage == dictx['GAGE']:
				gageSeperatedList.append(dictx)

			else:
				allList.append(gageSeperatedList)
				previousGage = dictx['GAGE']
				gageSeperatedList = []
				gageSeperatedList.append(dictx)

		#captures the last list that did not get appende in if previousGage == dictx['GAGE']:
		allList.append(gageSeperatedList)
		orderedList = []
		for listx in allList:
			#Sorts the lists that contain the same gage by the drawing number
			listSortedByDrawingNumber = sorted(listx, key=itemgetter('DRAWING'))
			for x in listSortedByDrawingNumber:
				orderedList.append(x)

		return orderedList

	#DO not sort if the file is a burn.
	else:
		return excelInfo



class mainUIClass:

	def __init__(self,master):
		self.master = master
		self.master.title("PDF printer")
		self.selectedPrinter = StringVar()
		self.selectedPaper = StringVar()
		self.selectedPaper.set('letter')
		self.jobCounter = 0
		self.PDFs = None
		self.user = os.getlogin()
		self.POPENFile = 'C:\\Users\\'+self.user+'\\tmpgs'
		self.print_wrong_revision_var = IntVar()
		self.current_window = master
		global isFrozen
		self.runningInFrozen = isFrozen
		"""
		sizex = 600
		sizey = 400
		posx  = 500
		posy  = 400
		master.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
		"""

		self.top_label = Label (master, text="Browse for the spreadsheet that contains the partnumbers.")
		self.top_label.grid(row=0, column=1, sticky=W)


		self.entryVar = StringVar()
		#self.entryVar.set("Browse for burn file")
		self.path_show = Label(master, width=100, background="white", anchor=W, relief=GROOVE, height=1, textvariable=self.entryVar)
		self.path_show.grid(row=1, column=1,columnspan=2, padx=5)

		self.browse_button = Button(master, text="Browse", command=self.askFilename)
		self.browse_button.grid(row=1, column=3, sticky=E, padx=10)


		#Headers
		self.label_found_files_header = Label(master,text="Files that were found:")
		self.label_found_files_header.grid(row=3, column=1,sticky=W)

		self.label_unfound_files_header = Label(master,text="Parts that were NOT found:")
		self.label_unfound_files_header.grid(row=3, column=2,sticky=W)

		self.label_wrong_revision_header = Label(master,text="Parts with different rev level in spreadsheet:")
		self.label_wrong_revision_header.grid(row=5, column=2,sticky=W)

		#Bodys
		self.text_found_files_body = Text(master, background="white",height=40, width=60, relief=GROOVE)
		self.text_found_files_body.grid(row=4, column=1, sticky=E, rowspan=4, padx=3)

		self.text_unfound_files_body = Text(master, background="white",height=19, width=40, relief=GROOVE)
		self.text_unfound_files_body.grid(row=4, column=2, sticky=E, padx=3,columnspan=2)

		self.text_revision_body = Text(master, background="white",height=19, width=40, relief=GROOVE)
		self.text_revision_body.grid(row=6, column=2, sticky=E,padx=3,columnspan=2)

		self.text_found_files_body.configure(state=DISABLED)
		self.text_unfound_files_body.configure(state=DISABLED)
		self.text_revision_body.configure(state=DISABLED)

		#Progress bar
		#self.progress = ttk.Progressbar(master, orient="horizontal",length=400, mode="determinate")
		#self.progress.grid(row=14,column=1, columnspan=4, pady=20,sticky=W+E, padx=10)

		#Buttons
		self.close_button = Button(master, text="Close", command=master.quit)
		self.close_button.grid(row=15, column=1,sticky=W,pady=5,padx=4)

		self.options_button = Button(master, text="Printer Settings",command=self.create_options_window)
		self.options_button.grid(row=15,column=1,pady=5,sticky=E,padx=10)

		self.print_wrong_revision_checkbox = Checkbutton(master, text="Print different revision files", variable=self.print_wrong_revision_var)
		self.print_wrong_revision_checkbox.grid(row=15, column=2,sticky=E)


		self.excecute_button = Button(master, text="Print", command=self.checkSettingsBeforePrint)
		self.excecute_button.grid(row=15,column=3,sticky=E,pady=5,padx=10)


	def checkSettingsBeforePrint(self):

		if self.selectedPrinter.get() != '' and self.PDFs != None :
			printingThread = threading.Thread(target=self.printFiles)
			printingThread.start()
		else:
			if self.selectedPrinter.get() == '':
				posx  = 500
				posy  = 400
				sizex = 500
				sizey = 100
				top = Toplevel()
				top.grid_rowconfigure(0,weigh=1)
				top.grid_columnconfigure(0, weight=1)
				top.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
				top.title("Printer not set")
				msg = Message(top, text="Set the default printer in\nPrinter Settings.",width=200, pady=10)
				msg.grid(row=0, column=0,columnspan=5)
				button = Button(top,text="Ok", command=top.destroy)
				button.grid(row=1, column=0)
				self.current_window = top
				if self.runningInFrozen:
					top.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
				else:
					top.iconbitmap("emblem_print.ico")
				top.focus_force()
				top.bind("<FocusOut>", self.Alarm)
				return None
			elif self.PDFs == None:
				posx  = 500
				posy  = 400
				sizex = 500
				sizey = 100
				top = Toplevel()
				top.grid_rowconfigure(0,weigh=1)
				top.grid_columnconfigure(0, weight=1)
				top.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
				top.title("No file loaded")
				msg = Message(top, text="Browse for a file before printing.",width=200, pady=10)
				msg.grid(row=0, column=0,columnspan=5)
				button = Button(top,text="Ok", command=top.destroy)
				button.grid(row=1, column=0)
				self.current_window = top
				if self.runningInFrozen:
					top.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
				else:
					top.iconbitmap("emblem_print.ico")
				top.focus_force()
				top.bind("<FocusOut>", self.Alarm)
				return None

	def askFilename(self):
		currdir = os.getcwd()
		filey = None
		filey = filedialog.askopenfilename(parent=self.master, initialdir=currdir, title='Select burn file')

		if type(filey) == str and filey != '':
			self.entryVar.set(filey)
			try:
				if not (str(self.entryVar.get()).split(".")[1] == "xls" or str(self.entryVar.get()).split(".")[1] == "XLS" or str(self.entryVar.get()).split(".")[1] == "xlsx" or str(self.entryVar.get()).split(".")[1] == "XLSX"):
					posx  = 500
					posy  = 400
					sizex = 500
					sizey = 100
					top = Toplevel()
					top.title("Wrong file type")
					top.grid_rowconfigure(0,weigh=1)
					top.grid_columnconfigure(0, weight=1)
					top.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
					msg = Message(top, text="You selected a wrong file type.\nPlease use xls or xlsx.", width=300, anchor=CENTER)
					msg.grid(row=0, column=0)
					button = Button(top,text="Ok", command=top.destroy)
					button.grid(row=1, column=0)
					self.entryVar.set("")
					self.current_window = top
					if self.runningInFrozen:
						top.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
					else:
						top.iconbitmap("emblem_print.ico")
					top.focus_force()
					top.bind("<FocusOut>", self.Alarm)
					return None
			except:
				posx  = 500
				posy  = 400
				sizex = 500
				sizey = 100
				top = Toplevel()
				top.title("Wrong file type")
				top.grid_rowconfigure(0,weigh=1)
				top.grid_columnconfigure(0, weight=1)
				top.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
				msg = Message(top, text="You selected a wrong file type.\nPlease use xls or xlsx.", width=300, anchor=CENTER)
				msg.grid(row=0, column=0)
				button = Button(top,text="Ok", command=top.destroy)
				button.grid(row=1, column=0)
				self.entryVar.set("")
				top.focus_force()
				if self.runningInFrozen:
					top.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
				else:
					top.iconbitmap("emblem_print.ico")
				self.current_window = top
				top.bind("<FocusOut>", self.Alarm)
				return None
		else:
			return None


		temp = readExcel(str(self.entryVar.get()))
		orderedExcelInfo = sortPartNumberList(temp[0],temp[1])
		temp = findPDFs(orderedExcelInfo)

		#The xlsx library is not reading the data correctly.
		#for dictx in orderedExcelInfo:
		#	print (dictx)


		self.PDFs = temp [0]
		self.unfoundItems = temp[1]
		self.wrongRevsion = temp[2]

		#Populate the text fields
		self.text_found_files_body.configure(state=NORMAL)
		self.text_unfound_files_body.configure(state=NORMAL)
		self.text_revision_body.configure(state=NORMAL)

		self.text_found_files_body.delete('0.0', END)
		self.text_unfound_files_body.delete('0.0', END)
		self.text_revision_body.delete('0.0', END)


		counter = 1
		for pdf in self.PDFs:
			self.text_found_files_body.insert(str(counter)+'.0',str(counter)+". "+pdf.replace("\"","")+"\n")
			counter += 1
		#print (counter)

		self.totalFiles = counter
		counter = 1

		for part in self.unfoundItems:
			self.text_unfound_files_body.insert(str(counter)+'.0', str(counter)+". "+part.replace("\"","")+"\n")
			counter += 1

		counter = 1
		for part in self.wrongRevsion:
			self.text_revision_body.insert(str(counter)+'.0',str(counter)+". "+part.replace("\"","")+"\n")
			counter += 1

		self.text_found_files_body.configure(state=DISABLED)
		self.text_unfound_files_body.configure(state=DISABLED)
		self.text_revision_body.configure(state=DISABLED)


	def Alarm(self, event):
		self.current_window.focus_force()
		self.current_window.bell()

	def create_options_window(self):
		listx = []

		for printerDetails in win32print.EnumPrinters(2):
			listx.append(printerDetails[2])
		self.printerList = listx

		self.options_windows = tkinter.Toplevel(self.master)
		self.options_windows.title("Options")

		if self.runningInFrozen:
			self.options_windows.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
		else:
			self.options_windows.iconbitmap("emblem_print.ico")

		counter = 1
		if self.selectedPrinter.get() == '':
			self.selectedPrinter.set(self.printerList[0])

		self.options_label_printer = Label(self.options_windows,text="Select Printer")
		self.options_label_printer.grid(row=0,column=0, sticky=W)

		self.options_label_paper = Label(self.options_windows,text="Select Paper")
		self.options_label_paper.grid(row=0,column=1, sticky=W)

		for printer in self.printerList:
			tempRadio = Radiobutton(self.options_windows, text=printer, variable=self.selectedPrinter, value=printer)
			tempRadio.grid(row=counter, column=0, sticky=W)
			counter += 1

		self.selectedPaper.set('letter')
		self.letter_radio = Radiobutton(self.options_windows, text='8 1/2x11', variable=self.selectedPaper, value='letter')
		self.letter_radio.grid(row=1,column=1,sticky=W)
		self.ledger_radio = Radiobutton(self.options_windows, text='11x17', variable=self.selectedPaper, value='ledger')
		self.ledger_radio.grid(row=2,column=1,sticky=W)


		local_close_button = Button(self.options_windows, text="Ok", command=self.save_options_and_destroy_options_window)
		local_close_button.grid(column=0,row=counter)

		self.current_window = self.options_windows
		self.options_windows.focus_force()
		self.options_windows.bind("<FocusOut>", self.Alarm)

	def save_options_and_destroy_options_window(self):
		f = open('C:\\Users\\'+self.user+'\\pdf_printer_settings','w+')
		f.write(self.selectedPrinter.get())
		f.close()
		self.options_windows.destroy()


	def printFiles(self):

		for pdf in self.PDFs:
			self.jobCounter = ghostscript("\"T:\RELEASED_FILES\CURRENT_PDF\\"+pdf.replace("\"","")+"\"", self.jobCounter, self.selectedPrinter.get(), self.selectedPaper.get(),self.POPENFile)
			#print (jobCounter)

		if self.print_wrong_revision_var.get():
			for pdf in self.wrongRevsion:
				self.jobCounter = ghostscript("\"T:\RELEASED_FILES\CURRENT_PDF\\"+pdf.replace("\"","")+"\"", self.jobCounter, self.selectedPrinter.get(), self.selectedPaper.get(),self.POPENFile)


		#os.remove(self.POPENFile)
		posx  = 500
		posy  = 400
		sizex = 500
		sizey = 100
		top = Toplevel()
		top.title("Done")
		top.grid_rowconfigure(0,weigh=1)
		top.grid_columnconfigure(0, weight=1)
		top.wm_geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
		msg = Message(top, text="Sent all files to printer.\nPlease wait for the printer to finish", width=200, pady=10)
		msg.grid(row=0, column=0, columnspan=4)
		top.focus_force()
		self.current_window = top
		if self.runningInFrozen:
			top.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
		else:
			top.iconbitmap("emblem_print.ico")
		top.bind("<FocusOut>", self.Alarm)
		button = Button(top,text="Ok", command=top.quit)
		button.grid(row=1, column=0)

def ghostscript(pdfPath, jobCounter, printer,paperType,filex):
	filex = open(filex,'a')
	filex.write(asctime()+"\n")
	filex.write(str(jobCounter)+"\n")
	filex.write(printer+"\n")
	filex.write(pdfPath+"\n")
	filex.write(paperType+"\n")
	#filex.write(str(os.getcwd())+"\n")
	global isFrozen
	if isFrozen:
		if paperType =='letter':
			command = sys._MEIPASS+r"\gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
			#command = r"gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
			#the following string shoudl generate something like this: -sOutputFile="\\spool\KONICA MINOLTA 423"
			command = command + "\"\\\\spool\\"+printer+"\" "
			command = command + pdfPath
		else:
			command = sys._MEIPASS+r"\gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
			#command = r"gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
			#the following string shoudl generate something like this: -sOutputFile="\\spool\KONICA MINOLTA 423"
			command = command + "\"\\\\spool\\"+printer+"\" "
			command = command + pdfPath
		#print (command)
	else:
		command = r"gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
		#command = r"gswin64c.exe -dPrinted -q -sDEVICE=mswinpr2 -dNoCancel -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE -sOutputFile="
		#the following string shoudl generate something like this: -sOutputFile="\\spool\KONICA MINOLTA 423"
		command = command + "\"\\\\spool\\"+printer+"\" "
		command = command + pdfPath
	"""
	command = r"gswin64c.exe -dPrinted -q -dNoCancel -sPAPERSIZE="+paperType+" -dBATCH -dFitPage -dNOPROMPT -dFIXEDMEDIA -dNOPAUSE "
	command = command + pdfPath
	"""
	filex.write(command+"\n")

	#print (command)


	try:
		gs = subprocess.Popen(command, shell=True, stdout=filex,stderr=subprocess.STDOUT,stdin=subprocess.PIPE)
	except:
		filex.write("POPEN\n")
		filex.write(str(sys.exc_info()[0]))
		filex.write(str(sys.exc_info()[1]))
		filex.write("\n")
	try:
		while True:
			gs.poll()
			if gs.returncode != None:
				#print ("Finished")
				return jobCounter + 1
	except:
		filex.write("returncode\n")
		filex.write(str(sys.exc_info()[0]))
		filex.write(str(sys.exc_info()[1]))
		filex.write("\n")


	filex.write("\n\n")
	f.close()

if __name__=="__main__":
	global isFrozen

	f = open('C:\\Users\\'+os.getlogin()+'\\tmpgs', 'w+')
	f.truncate(0)
	f.close()

	#checkGhostScriptPath()
	root = tkinter.Tk()
	if getattr(sys, 'frozen', False):
		isFrozen = True
		root.iconbitmap(sys._MEIPASS+r"/emblem_print.ico")
	else:
		root.iconbitmap("emblem_print.ico")
		isFrozen = False
	root.style = ttk.Style()
	mainUI = mainUIClass(root)
	root.style.theme_use("winnative")

	try:
		f = open('C:\\Users\\'+os.getlogin()+'\\pdf_printer_settings','r+')
	except IOError:
		f = open('C:\\Users\\'+os.getlogin()+'\\pdf_printer_settings','w+')

	mainUI.selectedPrinter.set(f.readline().rstrip())
	f.close()
	root.mainloop()
