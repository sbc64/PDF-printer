import os, sys
import win32print
from xlrd import open_workbook
from openpyxl import load_workbook
import subprocess
from tkinter import filedialog, Tk, Label, Button, LEFT, RIGHT, W, Message, StringVar, Toplevel, Listbox, ttk, Text
from tkinter import *
import tkinter
from tkinter import ttk
import threading
from operator import itemgetter
import io
from time import asctime
#import win32api
import winreg

def determineColumnNumbers(fileToRead):


	return (partColumn, drawingColumn, revColumn, gageColumn)

def readExcel(fileToRead):
	partNumberList, drawingNOList, revLevelList, gageList = [],[],[],[]
	excelTempDict = {'PARTNO':'','DRAWING':'','REV':'','GAGE':''}
	dictList = []
	if fileToRead:
		if fileToRead.split(".")[1] == "xls" or fileToRead.split(".")[1] == "XLS":
			wb = open_workbook(str(fileToRead))
			ws = wb.sheets()[0]
			for row in range(ws.nrows):
				excelTempDict = {}
				partnumber = str(ws.cell(row,0).value).rstrip()
				if not partnumber == 'partno':
					if (partnumber[-1] == 'F'):
						excelTempDict['PARTNO']=partnumber[:-1]
					elif(partnumber[-1] == 'S'):
						excelTempDict['PARTNO']=partnumber
					else:
						excelTempDict['PARTNO']=partnumber

					excelTempDict['DRAWING']=str(ws.cell(row,4).value).rstrip()
					excelTempDict['REV']=str(ws.cell(row,5).value).rstrip()
					excelTempDict['GAGE']=str(ws.cell(row,8).value).rstrip()
					dictList.append(excelTempDict)

		elif fileToRead.split(".")[1] == "xlsx" or fileToRead.split(".")[1] == "XLSX":
			ws = load_workbook(filename=str(fileToRead))['Sheet1']
			for row in ws.rows:
				excelTempDict = {}
				partnumber = str(row[0].value).rstrip()
				if not partnumber == 'partno':
					if (partnumber[-1] == 'F'):
						excelTempDict['PARTNO']=partnumber[:-1]
					elif(partnumber[-1] == 'S'):
						excelTempDict['PARTNO']=partnumber
					else:
						excelTempDict['PARTNO']=partnumber

					excelTempDict['DRAWING']=str(row[4].value).rstrip()
					excelTempDict['REV']=str(row[5].value).rstrip()
					excelTempDict['GAGE']=str(row[8].value).rstrip()
					dictList.append(excelTempDict)
	return dictList



if __name__=="__main__":

	readExcel()
